
import re
import json
from typing import List, Dict, Any, Optional, Tuple
import openpyxl
class TableToTreeConverter:
    """通用表格转层次化字典转换器"""
    
    def __init__(self):
        pass
    
    @staticmethod
    def is_likely_numeric(s):
        """判断一个字符串是否可能是数字或财务符号"""
        if not isinstance(s, str):
            return isinstance(s, (int, float))
        s = s.strip().replace(',', '').replace('$', '').replace('￥', '').replace('%', '')
        return bool(re.fullmatch(r'[\d\.]+|\([\d\.]+\)|-', s)) or s == ''
    
    @staticmethod
    def clean_value(value):
        """清理和转换数值"""
        if isinstance(value, (int, float)):
            return value
        if isinstance(value, str):
            value = value.strip()
            if value == '' or value == '-':
                return value
            # 处理括号表示的负数
            if value.startswith('(') and value.endswith(')'):
                try:
                    return value  # 保持原格式
                except:
                    return value
            try:
                # 尝试转换为整数
                return int(value.replace(',', ''))
            except (ValueError, TypeError):
                try:
                    # 尝试转换为浮点数
                    return float(value.replace(',', ''))
                except (ValueError, TypeError):
                    return value
        return value
    
    def get_leaf_columns(self, col_h: Dict[str, List[str]]) -> List[str]:
        """从列层次结构中提取叶子节点列标题"""
        leaf_columns = []
        
        # 获取所有父节点和子节点
        all_children = set(child for children in col_h.values() for child in children)
        
        # 按顺序遍历并提取叶子节点
        for parent, children in col_h.items():
            if not children:  # 父节点没有子节点，它本身就是叶子
                if parent not in leaf_columns:
                    leaf_columns.append(parent)
            else:  # 有子节点，子节点就是叶子
                for child in children:
                    if child not in leaf_columns:
                        leaf_columns.append(child)
        
        return leaf_columns
    
    def find_data_start_column(self, table: List[List], column_headers: List[str]) -> int:
        """找到数据开始的列索引"""
        for row in table:
            try:
                return row.index(column_headers[0])
            except (ValueError, IndexError):
                continue
        # 默认从第1列开始（索引1）
        return 1
    
    def create_row_lookup(self, table: List[List]) -> Dict[str, List]:
        """创建行标题到行数据的映射"""
        return {str(row[0]).strip(): row for row in table if row and row[0]}
    
    def build_nested_structure(self, row_lookup: Dict[str, List], row_h: Dict[str, List[str]], 
                              column_headers: List[str], data_start_col: int) -> Dict[str, Any]:
        """递归构建嵌套字典结构"""
        
        def build_level(parent_data: Dict, keys_to_process: List[str]):
            for key in keys_to_process:
                # 情况A: key是父节点，有子节点
                if key in row_h and row_h[key]:
                    parent_data[key] = {}
                    build_level(parent_data[key], row_h[key])
                # 情况B: key是叶子节点，对应一行数据
                elif key in row_lookup:
                    data_row = row_lookup[key]
                    # 提取数据值
                    data_values = data_row[data_start_col : data_start_col + len(column_headers)]
                    
                    # 创建列名到值的映射
                    parent_data[key] = {
                        col_name: self.clean_value(data_values[i]) if i < len(data_values) else None
                        for i, col_name in enumerate(column_headers)
                    }
        
        # 找到顶级父节点（不是任何其他节点的子节点）
        all_children = set(child for children in row_h.values() for child in children)
        top_level_keys = [key for key in row_h.keys() if key not in all_children]
        
        result = {}
        build_level(result, top_level_keys)
        return result
    
    def convert_table(self, table: List[List], row_h: Dict[str, List[str]], 
                     col_h: Dict[str, List[str]]) -> Dict[str, Any]:
        """
        将表格转换为层次化字典
        
        Args:
            table: 表格数据 (List[List] 格式)
            row_h: 行层次结构字典
            col_h: 列层次结构字典
            
        Returns:
            层次化的字典结构
        """
        
        # 1. 获取叶子列标题
        column_headers = self.get_leaf_columns(col_h)
        if not column_headers:
            raise ValueError("无法从列层次结构中提取有效的列标题")
        
        # 2. 找到数据开始的列索引
        data_start_col = self.find_data_start_column(table, column_headers)
        
        # 3. 创建行查找映射
        row_lookup = self.create_row_lookup(table)
        
        # 4. 构建嵌套结构
        result = self.build_nested_structure(row_lookup, row_h, column_headers, data_start_col)
        
        return result
class mytableconverter:
    """通用表格转层次化字典转换器 (修复版)"""
    def __init__(self):
        pass

    @staticmethod
    def is_likely_numeric(s):
        """判断一个字符串是否可能是数字或财务符号"""
        if not isinstance(s, str):
            return isinstance(s, (int, float))
        s = s.strip().replace(',', '').replace('$', '').replace('￥', '').replace('%', '')
        return bool(re.fullmatch(r'[\d\.]+|\([\d\.]+\)|-', s)) or s == ''

    @staticmethod
    def clean_value(value):
        """清理和转换数值"""
        if isinstance(value, (int, float)):
            return value
        if isinstance(value, str):
            value = value.strip()
            if value == '' or value == '–': # 在第二个例子中是 '–' 而不是 '-'
                return value
            # 处理括号表示的负数
            if value.startswith('(') and value.endswith(')'):
                return value  # 保持原格式
            try:
                # 尝试转换为整数
                return int(value.replace(',', ''))
            except (ValueError, TypeError):
                try:
                    # 尝试转换为浮点数
                    return float(value.replace(',', ''))
                except (ValueError, TypeError):
                    return value
        return value

    def _get_full_column_headers(self, col_h: Dict[str, List[str]]) -> List[Tuple[str, ...]]:
        """
        从列层次结构中提取每个叶子节点的完整路径。
        例如：{'2019': ['$', '%']} -> [('2019', '$'), ('2019', '%')]
        """
        full_headers = []
        # 假设 col_h 的键（父节点）是有序的
        for parent, children in col_h.items():
            if children:
                for child in children:
                    full_headers.append((parent, child))
            else:
                # 如果父节点没有子节点，它本身就是完整路径
                full_headers.append((parent,))
        return full_headers

    def find_data_start_column(self, table: List[List], column_headers: List[str]) -> int:
        """找到数据开始的列索引"""
        # 使用叶子节点列来定位
        leaf_columns = [h[-1] for h in column_headers]
        if not leaf_columns:
            return 1 # 默认值

        for row in table:
            try:
                # 尝试找到第一个叶子节点列标题的位置
                return row.index(leaf_columns[0])
            except (ValueError, IndexError):
                continue
        # 默认从第1列开始（索引1）
        return 1

    def create_row_lookup(self, table: List[List], header_rows_count: int) -> Dict[str, List]:
        """创建行标题到行数据的映射，跳过表头行"""
        # 一个简单的启发式方法：表头之后是数据行
        data_rows = table[header_rows_count:]
        return {str(row[0]).strip(): row for row in data_rows if row and str(row[0]).strip()}

    def build_nested_structure(self, row_lookup: Dict[str, List], row_h: Dict[str, List[str]],
                               column_headers: List[Tuple[str, ...]], data_start_col: int) -> Dict[str, Any]:
        """递归构建嵌套字典结构"""

        def build_level(parent_data: Dict, keys_to_process: List[str]):
            for key in keys_to_process:
                # 情况A: key是父节点，有子节点
                if key in row_h and row_h.get(key):
                    parent_data[key] = {}
                    build_level(parent_data[key], row_h[key])
                # 情况B: key是叶子节点，对应一行数据
                elif key in row_lookup:
                    data_row = row_lookup[key]
                    data_values = data_row[data_start_col:]
                    
                    # 为这一行创建一个嵌套字典
                    row_data_nested = {}
                    
                    # 遍历每个列的完整路径和对应的数据值
                    for i, header_path in enumerate(column_headers):
                        if i >= len(data_values):
                            continue
                        
                        value = self.clean_value(data_values[i])
                        
                        # 使用一个指针来构建嵌套字典
                        current_level = row_data_nested
                        # 遍历路径中的所有部分，除了最后一个
                        for part in header_path[:-1]:
                            # 如果键不存在，则创建一个新字典
                            current_level = current_level.setdefault(part, {})
                        
                        # 在最深层级设置值
                        current_level[header_path[-1]] = value
                        
                    parent_data[key] = row_data_nested

        # 找到顶级父节点（不是任何其他节点的子节点）
        all_children = set(child for children in row_h.values() for child in children)
        top_level_keys = [key for key in row_h.keys() if key not in all_children]

        result = {}
        build_level(result, top_level_keys)
        return result

    def convert(self, original_table: List[List[str]], row_h: Dict[str, List[str]], col_h: Dict[str, List[str]]) -> Dict[str, Any]:
        """
        执行转换的主方法。
        """
        # 1. 获取完整的、带层级的列标题
        full_column_headers = self._get_full_column_headers(col_h)
        
        # 2. 找到数据列的起始索引
        # 简单的启发式：计算可能的表头行数。这里假设表头是直到第一个数据行标题出现前的所有行。
        first_data_row_key = list(row_h.keys())[0]
        header_rows_count = 0
        for i, row in enumerate(original_table):
            if row and str(row[0]).strip() == first_data_row_key:
                header_rows_count = i
                break
        
        leaf_headers_for_search = [h[-1] for h in full_column_headers]
        data_start_col = self.find_data_start_column(original_table[:header_rows_count+1], leaf_headers_for_search)

        # 3. 创建行标题到数据的查找表
        row_lookup = self.create_row_lookup(original_table, header_rows_count)

        # 4. 构建最终的嵌套结构
        tree_table = self.build_nested_structure(row_lookup, row_h, full_column_headers, data_start_col)
        
        return tree_table
def extract_hierarchies(llm_output: str):
    """
    从LLM输出中提取行和列层次结构
    
    Args:
        llm_output (str): LLM的完整输出文本
        
    Returns:
        tuple: (row_hierarchy_json, column_hierarchy_json)
    """
    row_hierarchy = None
    column_hierarchy = None
    
    # 找到所有 ```json 代码块
    json_pattern = r'```json\s*(.*?)```'
    json_matches = re.findall(json_pattern, llm_output, re.DOTALL | re.IGNORECASE)
    
    # 找到所有 ```json 的位置
    json_positions = []
    for match in re.finditer(r'```json', llm_output, re.IGNORECASE):
        json_positions.append(match.start())
    
    # 遍历每个json块
    for i, json_text in enumerate(json_matches):
        try:
            # 解析JSON
            parsed_json = json.loads(json_text.strip())
            
            # 获取这个json块之前的文本（最多往前看500个字符）
            if i < len(json_positions):
                start_pos = max(0, json_positions[i] - 500)
                preceding_text = llm_output[start_pos:json_positions[i]].lower()
                
                # 检查是否是行层次结构
                if re.search(r'row', preceding_text) and row_hierarchy is None:
                    row_hierarchy = parsed_json
                    print(f"✅ 找到行层次结构 (第{i+1}个JSON块)")
                    
                # 检查是否是列层次结构  
                elif re.search(r'col', preceding_text) and column_hierarchy is None:
                    column_hierarchy = parsed_json
                    print(f"✅ 找到列层次结构 (第{i+1}个JSON块)")
                    
        except json.JSONDecodeError as e:
            print(f"⚠️ 第{i+1}个JSON块解析失败: {e}")
            continue
    
    return row_hierarchy, column_hierarchy

def normalize_tree_none_values(tree_table):
    """
    递归处理树形表格，将所有None值转换为字符串"None"
    
    Args:
        tree_table: 树形表格结构（可以是dict, list或其他类型）
        
    Returns:
        处理后的树形表格，所有None值都被替换为字符串"None"
    """
    if tree_table is None:
        return "None"
    
    if isinstance(tree_table, dict):
        # 递归处理字典的每个键值对
        normalized_dict = {}
        for key, value in tree_table.items():
            normalized_dict[key] = normalize_tree_none_values(value)
        return normalized_dict
    
    elif isinstance(tree_table, list):
        # 递归处理列表的每个元素
        return [normalize_tree_none_values(item) for item in tree_table]
    
    else:
        # 基本类型直接返回
        return tree_table

def print_tree(d, indent=0):
    """
    递归打印嵌套字典为树状结构
    :param d: dict 或基本类型
    :param indent: 缩进层级
    """
    for key, value in d.items():
        print('│   ' * indent + '├─ ' + str(key))
        if isinstance(value, dict):
            print_tree(value, indent + 1)
        else:
            print('│   ' * (indent + 1) + '├─ ' + str(value))
from model_clients import ModelClient, OpenaimodelClient
from prompt import GET_Hierarchy_Prompt_oai
# ------- 方案1 通过LLM提取树结构之后进行脚本提取转化为树 -------------
def table2tree_rule(table, model_name, temperature, model_type="oai"):
    """
    Input :
        table : 原始表格
        model_name : 模型名称
        temperature : 温度
        model_type : 模型类型 ("oai" 或 "opensource")
    Output :
        Tree_table : 树化表格
    """
    model = ModelClient()
    oaimodel = OpenaimodelClient(model = model_name)
    
    if model_type == "oai":
        response = oaimodel.generate(GET_Hierarchy_Prompt_oai.format(table=table), temperature=temperature)
        full_llm_output = response
    else:
        response = model.test_generate_stream_advanced(GET_Hierarchy_Prompt_oai.format(table=table), temperature=temperature)
        full_llm_output = "".join(list(response))
    
    print(full_llm_output)
    row_h, col_h = extract_hierarchies(full_llm_output)
    print(f"Row Hierarchy: {row_h}")
    print(f"Column Hierarchy: {col_h}")
    
    Tableconverter = mytableconverter()
    Tree_table = Tableconverter.convert(table, row_h, col_h)
    print(f"✅ 成功生成树形表格，根节点数量: {len(Tree_table)}")
    
    return Tree_table

# ------- 方案2 通过LLM进行三步操作提取树 -------------
def excel_to_markdown_with_cell_ref(sheet):
    """
    将Excel工作表转换为包含单元格位置的Markdown表格
    
    Args:
        sheet: openpyxl工作表对象
    
    Returns:
        str: Markdown格式的表格字符串
    """
    # 获取工作表的基本信息
    max_row = sheet.max_row
    max_col = sheet.max_column
    
    # 获取所有合并单元格信息
    merged_cells = {}
    for merged_range in sheet.merged_cells.ranges:
        # 将合并区域转换为字符串表示，如 "A1:B2"
        merged_ref = f"{get_column_letter(merged_range.min_col)}{merged_range.min_row}:{get_column_letter(merged_range.max_col)}{merged_range.max_row}"
        # 只记录左上角单元格的值
        top_left_cell = sheet.cell(merged_range.min_row, merged_range.min_col)
        merged_cells[merged_ref] = top_left_cell.value
    
    # 构建表格数据
    table_data = []
    for row in range(1, max_row + 1):
        row_data = []
        for col in range(1, max_col + 1):
            cell = sheet.cell(row, col)
            cell_ref = f"{get_column_letter(col)}{row}"
            
            # 检查当前单元格是否在某个合并区域内
            is_in_merged = False
            cell_value = None
            
            for merged_ref, value in merged_cells.items():
                start_ref, end_ref = merged_ref.split(':')
                start_col_str = ''.join(filter(str.isalpha, start_ref))
                start_row = int(''.join(filter(str.isdigit, start_ref)))
                end_col_str = ''.join(filter(str.isalpha, end_ref))
                end_row = int(''.join(filter(str.isdigit, end_ref)))
                
                start_col = openpyxl.utils.column_index_from_string(start_col_str)
                end_col = openpyxl.utils.column_index_from_string(end_col_str)
                
                if (start_row <= row <= end_row and start_col <= col <= end_col):
                    is_in_merged = True
                    # 如果是合并区域的左上角单元格
                    if row == start_row and col == start_col:
                        cell_value = f"{merged_ref} {value if value is not None else ''}"
                    break
            
            if not is_in_merged:
                cell_value = f"{cell_ref} {cell.value if cell.value is not None else ''}"
            
            row_data.append(cell_value if cell_value is not None else f"{cell_ref} ")
        
        table_data.append(row_data)
    
    # 生成Markdown表格
    markdown_lines = []
    
    # 表头行
    header_row = "| " + " | ".join(str(table_data[0][col]) for col in range(max_col)) + " |"
    markdown_lines.append(header_row)
    
    # 分隔线行
    separator_row = "| " + " | ".join(["---"] * max_col) + " |"
    markdown_lines.append(separator_row)
    
    # 数据行
    for row in range(1, len(table_data)):
        data_row = "| " + " | ".join(str(table_data[row][col]) for col in range(max_col)) + " |"
        markdown_lines.append(data_row)
    
    return "\n".join(markdown_lines)

from prompt import HEADER_ANALYSIS_PROMPT, HIERARCHY_VALUE_IDENTIFICATION_PROMPT, FINAL_JSON_TREE_CONSTRUCTION
from prompt import HEADER_ANALYSIS_PROMPT_ENHANCED, HIERARCHY_VALUE_IDENTIFICATION_PROMPT_ENHANCED, FINAL_JSON_TREE_CONSTRUCTION_ENHANCED
class llm_based_converter:
    def __init__(self, model_type, model_name, api_key="", base_url=""):
        self.model = ModelClient()
        self.oaimodel = OpenaimodelClient(model=model_name, api_key=api_key, base_url=base_url)
        self.model_type = model_type
        self.if_print_prompt = True

    def extract_response(self, response):
        return response.split("```json")[1].split("```")[0]

    def extract_after_think(self, response):
        """提取</think>标签之后的内容"""
        if "</think>" in response:
            think_end = response.find("</think>")
            if think_end != -1:
                content_after_think = response[think_end + len("</think>"):].strip()
                return content_after_think
        return response.strip()

    def header_analysis_norm_enhance(self, table, temperature):
        if self.if_print_prompt:
            print("**Header Analysis Enhance PROMPT**")
            print(HEADER_ANALYSIS_PROMPT_ENHANCED.format(TABLE_AS_JSON_STRING=table))
        if self.model_type == "oai":
            response = self.oaimodel.generate(HEADER_ANALYSIS_PROMPT_ENHANCED.format(TABLE_AS_JSON_STRING=table), temperature=temperature)
            full_llm_output = response
            print("---Header Analysis Enhance---")
            print(full_llm_output)
        else:
            response = self.model.test_generate_stream(HEADER_ANALYSIS_PROMPT_ENHANCED.format(TABLE_AS_JSON_STRING=table), temperature=temperature)
            full_llm_output = "".join(list(response)) 
            full_llm_output = self.extract_after_think(full_llm_output)
        return full_llm_output

    def hierarchy_definition_enhance(self, table, Normalized_headers, temperature):
        if self.if_print_prompt:
            print("**Hierarchy Definition Enhance PROMPT**")
            print(HIERARCHY_VALUE_IDENTIFICATION_PROMPT_ENHANCED.format(TABLE_AS_JSON_STRING=table, NORMALIZED_HEADERS_FROM_STEP_1=Normalized_headers))
        if self.model_type == "oai":
            response = self.oaimodel.generate(HIERARCHY_VALUE_IDENTIFICATION_PROMPT_ENHANCED.format(TABLE_AS_JSON_STRING=table, NORMALIZED_HEADERS_FROM_STEP_1=Normalized_headers), temperature=temperature)
            full_llm_output = response
            print("---Hierarchy Definition Enhance---")
            print(full_llm_output)
        else:
            response = self.model.test_generate_stream(HIERARCHY_VALUE_IDENTIFICATION_PROMPT_ENHANCED.format(TABLE_AS_JSON_STRING=table, NORMALIZED_HEADERS_FROM_STEP_1=Normalized_headers), temperature=temperature)
            full_llm_output = "".join(list(response)) 
            full_llm_output = self.extract_after_think(full_llm_output)
        return full_llm_output

    def final_json_tree_construction_enhance(self, table, Normalized_headers, Hierarchy_definition, temperature):
        if self.if_print_prompt:
            print("**Final JSON Tree Construction Enhance PROMPT**")
            print(FINAL_JSON_TREE_CONSTRUCTION_ENHANCED.format(TABLE_AS_JSON_STRING=table, NORMALIZED_HEADERS_FROM_STEP_1=Normalized_headers, HIERARCHY_DEFINITION_FROM_STEP_2=Hierarchy_definition))
        if self.model_type == "oai":
            response = self.oaimodel.generate(FINAL_JSON_TREE_CONSTRUCTION_ENHANCED.format(TABLE_AS_JSON_STRING=table, NORMALIZED_HEADERS_FROM_STEP_1=Normalized_headers, HIERARCHY_DEFINITION_FROM_STEP_2=Hierarchy_definition), temperature=temperature)
            full_llm_output = response
            print("---Final JSON Tree Construction Enhance---")
            print(full_llm_output)
        else:
            response = self.model.test_generate_stream(FINAL_JSON_TREE_CONSTRUCTION_ENHANCED.format(TABLE_AS_JSON_STRING=table, NORMALIZED_HEADERS_FROM_STEP_1=Normalized_headers, HIERARCHY_DEFINITION_FROM_STEP_2=Hierarchy_definition), temperature=temperature)
            full_llm_output = "".join(list(response)) 
            full_llm_output = self.extract_after_think(full_llm_output)
        return full_llm_output

    def header_analysis_norm(self, table, temperature):
        if self.if_print_prompt:
            print("**Header Analysis Normal PROMPT**")
            print(HEADER_ANALYSIS_PROMPT.format(TABLE_AS_JSON_STRING=table))
        if self.model_type == "oai":
            response = self.oaimodel.generate(HEADER_ANALYSIS_PROMPT.format(TABLE_AS_JSON_STRING=table), temperature=temperature)
            full_llm_output = response
            print("---Header Analysis Normal---")
            print(full_llm_output)
        else:
            response = self.model.test_generate_stream(HEADER_ANALYSIS_PROMPT.format(TABLE_AS_JSON_STRING=table), temperature=temperature)
            full_llm_output = "".join(list(response)) 
            full_llm_output = self.extract_after_think(full_llm_output)
        return full_llm_output
    
    def hierarchy_definition(self, table, Normalized_headers, temperature):
        if self.if_print_prompt:
            print("**Hierarchy Definition PROMPT**")
            print(HIERARCHY_VALUE_IDENTIFICATION_PROMPT.format(TABLE_AS_JSON_STRING=table, NORMALIZED_HEADERS_FROM_STEP_1=Normalized_headers))
        if self.model_type == "oai":
            response = self.oaimodel.generate(HIERARCHY_VALUE_IDENTIFICATION_PROMPT.format(TABLE_AS_JSON_STRING=table, NORMALIZED_HEADERS_FROM_STEP_1=Normalized_headers), temperature=temperature)
            full_llm_output = response
            print("---Hierarchy Definition---")
            print(full_llm_output)
        else:
            response = self.model.test_generate_stream(HIERARCHY_VALUE_IDENTIFICATION_PROMPT.format(TABLE_AS_JSON_STRING=table, NORMALIZED_HEADERS_FROM_STEP_1=Normalized_headers), temperature=temperature)
            full_llm_output = "".join(list(response)) 
            full_llm_output = self.extract_after_think(full_llm_output)
        return full_llm_output
    
    def final_json_tree_construction(self, table, Normalized_headers, Hierarchy_definition, temperature):
        if self.if_print_prompt:
            print("**Final JSON Tree Construction PROMPT**")
            print(FINAL_JSON_TREE_CONSTRUCTION.format(TABLE_AS_JSON_STRING=table, NORMALIZED_HEADERS_FROM_STEP_1=Normalized_headers, HIERARCHY_DEFINITION_FROM_STEP_2=Hierarchy_definition))
        if self.model_type == "oai":
            response = self.oaimodel.generate(FINAL_JSON_TREE_CONSTRUCTION.format(TABLE_AS_JSON_STRING=table, NORMALIZED_HEADERS_FROM_STEP_1=Normalized_headers, HIERARCHY_DEFINITION_FROM_STEP_2=Hierarchy_definition), temperature=temperature)
            full_llm_output = response
            print("---Final JSON Tree Construction---")
            print(full_llm_output)
        else:
            response = self.model.test_generate_stream(FINAL_JSON_TREE_CONSTRUCTION.format(TABLE_AS_JSON_STRING=table, NORMALIZED_HEADERS_FROM_STEP_1=Normalized_headers, HIERARCHY_DEFINITION_FROM_STEP_2=Hierarchy_definition), temperature=temperature)
            full_llm_output = "".join(list(response)) 
            full_llm_output = self.extract_after_think(full_llm_output)
        return full_llm_output

    def parse_json_with_merge(self, json_string):
        """
        解析包含重复key的JSON字符串，自动合并重复key
        """
        def merge_dicts(dict1, dict2):
            """递归合并两个字典"""
            result = dict1.copy()
            for key, value in dict2.items():
                if key in result:
                    if isinstance(result[key], dict) and isinstance(value, dict):
                        result[key] = merge_dicts(result[key], value)
                    # 如果不是字典，保留原值或选择其他策略
                else:
                    result[key] = value
            return result
        
        # 使用object_pairs_hook来处理重复key
        def object_pairs_hook(pairs):
            result = {}
            for key, value in pairs:
                if key in result:
                    # 如果key已存在，合并值
                    if isinstance(result[key], dict) and isinstance(value, dict):
                        result[key] = merge_dicts(result[key], value)
                    else:
                        # 其他策略：可以选择保留第一个、最后一个，或创建列表
                        pass
                else:
                    result[key] = value
            return result
        
        return json.loads(json_string, object_pairs_hook=object_pairs_hook)

    def parse_json_with_merge_oai(self, oai_output):
        """
        简单处理OpenAI输出的JSON解析函数
        """
        # 提取JSON内容
        if "```json" in oai_output:
            json_content = oai_output.split("```json")[1].split("```")[0].strip()
        elif "```" in oai_output:
            json_content = oai_output.split("```")[1].strip()
        else:
            # 如果没有代码块标记，查找{}之间的内容
            start = oai_output.find('{')
            end = oai_output.rfind('}') + 1
            json_content = oai_output[start:end].strip()
        
        # 处理重复key的hook函数
        def object_pairs_hook(pairs):
            result = {}
            for key, value in pairs:
                if key in result:
                    if isinstance(result[key], dict) and isinstance(value, dict):
                        # 合并字典
                        result[key].update(value)
                    else:
                        # 保留最后一个值
                        result[key] = value
                else:
                    result[key] = value
            return result
        def _deep_merge_dicts(dict1: dict, dict2: dict) -> dict:
            """
            递归地将 dict2 合并到 dict1 中。
            
            - 如果键只存在于一个字典中，则保留它。
            - 如果键在两个字典中都存在，并且值都是字典，则递归合并这两个子字典。
            - 如果键在两个字典中都存在，但值不是两个字典（例如，字符串、列表或一个字典一个列表），
            则将它们合并到一个列表中（以避免数据丢失）。
            """
            merged = dict1.copy()

            for key, value2 in dict2.items():
                if key in merged:
                    value1 = merged[key]
                    
                    # 检查我们是否可以递归合并
                    if isinstance(value1, dict) and isinstance(value2, dict):
                        merged[key] = _deep_merge_dicts(value1, value2)
                    
                    # 如果不能合并（例如，一个是 dict，另一个是 str，或者都是 str）
                    # 我们必须将它们放入一个列表中以保留所有数据。
                    else:
                        # 如果它还不是列表，请将其转换为列表
                        if not isinstance(value1, list):
                            merged[key] = [value1]
                        
                        # 如果 value2 本身是一个列表，则扩展它，否则附加它
                        if isinstance(value2, list):
                            merged[key].extend(value2)
                        else:
                            merged[key].append(value2)
                else:
                    # 键是全新的，直接添加
                    merged[key] = value2
                    
            return merged

        # ----------------------------------------------------------------------
        # 2. 你的新 object_pairs_hook
        # ----------------------------------------------------------------------
        def recursive_merge_hook(pairs: list) -> dict:
            """
            在 json.loads 期间调用的 Hook，用于执行深度合并。
            """
            result = {}
            for key, value in pairs:
                if key in result:
                    # 发现重复键！
                    # 检查我们是否可以深度合并
                    if isinstance(result[key], dict) and isinstance(value, dict):
                        result[key] = _deep_merge_dicts(result[key], value)
                    else:
                        # 无法深度合并（例如，一个是 dict，另一个是 str）
                        # 回退到列表策略以避免数据丢失
                        if not isinstance(result[key], list):
                            result[key] = [result[key]]
                        
                        if isinstance(value, list):
                            result[key].extend(value)
                        else:
                            result[key].append(value)
                else:
                    # 这是一个新键，只需添加它
                    result[key] = value
                    
            return result
        return json.loads(json_content, object_pairs_hook=recursive_merge_hook)

    # --------------------------根据规则根据前两步给出的headers和第一次层级进行构建 -------------------------------
    def merge_duplicate_keys(data_rows, hierarchy_keys):
        """
        处理重复key的情况，当出现重复时进行合并而不是替换
        """
        result = {}
        
        for row in data_rows:
            current_dict = result
            
            # 遍历层级键
            for i, key_name in enumerate(hierarchy_keys):
                # 构造语义化的key名称：[Header Name] [Cell Value]
                if row[i]:  # 如果值不为空
                    semantic_key = f"{key_name} {row[i]}"
                else:
                    semantic_key = f"{key_name} "
                
                # 如果这是最后一个层级键，需要处理record数据
                if i == len(hierarchy_keys) - 1:
                    if semantic_key not in current_dict:
                        current_dict[semantic_key] = {}
                    
                    # 这里处理record数据...
                    # （后续会完善）
                else:
                    # 中间层级，如果key不存在就创建
                    if semantic_key not in current_dict:
                        current_dict[semantic_key] = {}
                    current_dict = current_dict[semantic_key]
        
        return result

    # 更完整的合并函数
    def build_merged_tree(self, table_data, normalized_headers, hierarchy_def):
        """
        构建合并重复key的树形结构
        """
        hierarchy_keys = hierarchy_def["hierarchy_keys"]
        value_leaves = hierarchy_def["value_leaves"]
        semantic_groups = hierarchy_def.get("semantic_groups", {})
        
        result = {}
        
        # 跳过表头，从数据行开始处理
        data_rows = table_data[2:]  # 假设前两行是表头
        
        for row in data_rows:
            current_dict = result
            
            # 按层级键构建嵌套结构
            for i, key_name in enumerate(hierarchy_keys):
                key_idx = normalized_headers.index(key_name)
                cell_value = row[key_idx] if key_idx < len(row) else ""
                
                # 构造语义化的key
                if cell_value:
                    semantic_key = f"{key_name} {cell_value}"
                else:
                    semantic_key = f"{key_name} "
                
                # 如果key不存在，创建新的字典
                if semantic_key not in current_dict:
                    current_dict[semantic_key] = {}
                
                current_dict = current_dict[semantic_key]
            
            # 在最深层添加value_leaves数据
            for group_name, group_fields in semantic_groups.items():
                if group_name not in current_dict:
                    current_dict[group_name] = {}
                
                for field in group_fields:
                    if field in normalized_headers:
                        field_idx = normalized_headers.index(field)
                        if field_idx < len(row):
                            value = row[field_idx]
                            # 尝试转换为数字
                            try:
                                if '.' in value:
                                    value = float(value)
                                else:
                                    value = int(value)
                            except (ValueError, AttributeError):
                                pass  # 保持原始字符串
                            
                            current_dict[group_name][field] = value
        
        return result

import re
from copy import deepcopy
from openpyxl.utils.cell import range_boundaries

def is_cell_position(text):
    """
    支持可选的绝对引用符号$: 如 $A$1
    """
    return bool(re.match(r'^\$?[A-Z]+\$?[0-9]+$', text.strip().upper()))

def is_cell_range(text):
    """
    匹配区域引用: 如 A1:B9、$A$1:$C$3
    """
    return bool(re.match(r'^\$?[A-Z]+\$?[0-9]+:\$?[A-Z]+\$?[0-9]+$', text.strip().upper()))

def get_value_from_ref(ref, sheet, range_policy="merged_top_left"):
    """
    range_policy:
      - "merged_top_left": 若为合并区域或普通区域，均取左上角值（默认）
      - "matrix": 非合并区域返回二维列表
      - "flatten": 非合并区域返回一维扁平列表
    """
    ref_u = ref.strip().upper()

    # 单个单元格
    if is_cell_position(ref_u):
        try:
            return sheet[ref_u].value
        except Exception as e:
            print(f"警告: 无法读取单元格 {ref}: {e}")
            return ref

    # 区域引用
    if is_cell_range(ref_u):
        try:
            min_col, min_row, max_col, max_row = range_boundaries(ref_u)

            # 若该区域正好是合并区域，则返回左上角的值
            for mr in sheet.merged_cells.ranges:
                if range_boundaries(str(mr).upper()) == (min_col, min_row, max_col, max_row):
                    return sheet.cell(row=min_row, column=min_col).value

            # 非合并区域，按策略返回
            cells = sheet[ref_u]
            if range_policy == "matrix":
                if isinstance(cells, tuple) and cells and isinstance(cells[0], tuple):
                    return [[c.value for c in row] for row in cells]
                else:
                    return [c.value for c in cells]
            if range_policy == "flatten":
                if isinstance(cells, tuple) and cells and isinstance(cells[0], tuple):
                    return [c.value for row in cells for c in row]
                else:
                    return [c.value for c in cells]

            # 默认：左上角
            return sheet.cell(row=min_row, column=min_col).value

        except Exception as e:
            print(f"警告: 无法读取区域 {ref}: {e}")
            return ref

    # 非引用字符串原样返回
    return ref

def convert_cell_positions_to_values(data_dict, sheet, range_policy="merged_top_left"):
    """
    将叶子节点的 'A1' 或 'A1:C3' 等引用转换为实际值/区域值
    """
    result = deepcopy(data_dict)

    def recursive_convert(obj):
        if isinstance(obj, dict):
            for k, v in obj.items():
                obj[k] = recursive_convert(v)
            return obj
        elif isinstance(obj, str):
            return get_value_from_ref(obj, sheet, range_policy=range_policy)
        else:
            return obj

    return recursive_convert(result)

from openpyxl.utils import get_column_letter, range_boundaries

def table2tree_llm(sheet, table, model_name, temperature, model_type, table2_tree_mode):
    """
    Input :
        sheet : 原始表格的sheet
        table : 原始表格
        model_name : 模型名称
        temperature : 温度
        model_type : 使用开源还是闭源模型
        table2_tree_mode : 表格转树形结构模式 "normal" 或 "enhanced" enhanced用来解决 大型表格
    Output :
        Tree_table_v1, Tree_table_v2 : 树化表格 V1 LLM直接生成， V2 LLM生成树结构之后进行脚本提取转化为树
    """
    table2tree_llm = llm_based_converter(model_type, model_name)
    Tree_table_v1 = None
    Tree_table_v2 = None
    if table2_tree_mode.lower() == "normal":    
        Normalized_headers = table2tree_llm.header_analysis_norm(table, temperature)
        Hierarchy_definition = table2tree_llm.hierarchy_definition(table, Normalized_headers, temperature)
        Final_json_tree = table2tree_llm.final_json_tree_construction(table, Normalized_headers, Hierarchy_definition, temperature)
        if model_type == "oai":
            Tree_table_v1 = table2tree_llm.parse_json_with_merge_oai(Final_json_tree)
        else:
            Tree_table_v1 = table2tree_llm.parse_json_with_merge(Final_json_tree)
        # Tree_table_v2 = table2tree_llm.build_merged_tree(table, Normalized_headers, Hierarchy_definition)
        Tree_table_v2 = None
    elif table2_tree_mode.lower() == "enhanced":
        markdown_table = excel_to_markdown_with_cell_ref(sheet)
        Normalized_headers = table2tree_llm.header_analysis_norm_enhance(markdown_table, temperature)
        Hierarchy_definition = table2tree_llm.hierarchy_definition_enhance(markdown_table, Normalized_headers, temperature)
        Final_json_tree = table2tree_llm.final_json_tree_construction_enhance(markdown_table, Normalized_headers, Hierarchy_definition, temperature)
        if model_type == "oai":
            Tree_table_v1 = table2tree_llm.parse_json_with_merge_oai(Final_json_tree)
        else:
            Tree_table_v1 = table2tree_llm.parse_json_with_merge(Final_json_tree)
        Tree_table_v1 = convert_cell_positions_to_values(Tree_table_v1, sheet)
    return Tree_table_v1, Tree_table_v2
