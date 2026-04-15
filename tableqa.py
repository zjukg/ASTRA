## 加载表格数据
import argparse
import json
import re
import hashlib
import os
import time
import tiktoken
from astra_config import get_cache_root, get_dataset_root, get_record_root
from table2tree import table2tree_rule, table2tree_llm
from treeqa import TreeQA
from tree_quality_evaluator import evaluate_tree_quality
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell

class TreeTableCache:
    """树形表格缓存管理器"""
    
    def __init__(self, cache_dir=None):
        # 如果没有指定缓存目录，使用默认的绝对路径
        if cache_dir is None:
            cache_dir = get_cache_root() / "tree_tables"
        
        self.cache_dir = str(cache_dir)
        # 确保缓存目录存在
        os.makedirs(self.cache_dir, exist_ok=True)
    
    def _get_table_hash(self, table):
        """为表格生成唯一的hash标识"""
        # 将表格转为字符串并计算hash
        table_str = json.dumps(table, sort_keys=True, ensure_ascii=False, default = str) # 如果表格中有datetime 日期单元格，可能无法hash，将其转化为str
        return hashlib.md5(table_str.encode('utf-8')).hexdigest()
    
    def _get_cache_path(self, table):
        """获取缓存文件路径"""
        table_hash = self._get_table_hash(table)
        return os.path.join(self.cache_dir, f"tree_table_{table_hash}.json")
    
    def has_cached_tree(self, table):
        """检查表格是否已有缓存的树形结构"""
        cache_path = self._get_cache_path(table)
        return os.path.exists(cache_path)
    
    def save_tree_table(self, table, tree_table, row_h, col_h):
        """保存树形表格和层次结构到缓存"""
        cache_path = self._get_cache_path(table)
        
        cache_data = {
            "tree_table": tree_table,
            "row_hierarchy": row_h,
            "column_hierarchy": col_h,
            "original_table": table,
            "timestamp": __import__('time').time()
        }
        
        try:
            with open(cache_path, 'w', encoding='utf-8') as f:
                json.dump(cache_data, f, ensure_ascii=False, indent=2, default=str)
            print(f"✅ 树形表格已保存到缓存: {cache_path}")
            return True
        except Exception as e:
            print(f"❌ 保存缓存失败: {e}")
            return False
    
    def load_tree_table(self, table):
        """从缓存加载树形表格"""
        cache_path = self._get_cache_path(table)
        
        try:
            with open(cache_path, 'r', encoding='utf-8') as f:
                cache_data = json.load(f)
            
            print(f"✅ 从缓存加载树形表格: {cache_path}")
            return (cache_data["tree_table"], 
                   cache_data["row_hierarchy"], 
                   cache_data["column_hierarchy"])
        except Exception as e:
            print(f"❌ 加载缓存失败: {e}")
            return None, None, None

class Table2TreeRecorder:
    """表格到树形结构转换的记录器"""
    
    def __init__(self, record_dir=None, dataset_name="tatqa", start_idx=0, end_idx=0, enable_quality_eval=False):
        # 如果没有指定记录目录，使用默认的绝对路径
        if record_dir is None:
            record_dir = get_record_root()
        
        self.record_dir = str(record_dir)
        self.enable_quality_eval = enable_quality_eval  # 是否启用质量评估
        
        # 生成带时间戳和参数的文件名
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        filename = f"table2tree_record_{dataset_name}_{start_idx}-{end_idx}_{timestamp}.json"
        self.record_file = os.path.join(self.record_dir, filename)
        
        # 同样生成批量结果文件名
        batch_filename = f"batch_qa_results_{dataset_name}_{start_idx}-{end_idx}_{timestamp}.json"
        self.batch_result_file = os.path.join(self.record_dir, batch_filename)
        
        # 如果启用质量评估，也生成评估结果文件名
        if enable_quality_eval:
            eval_filename = f"quality_eval_{dataset_name}_{start_idx}-{end_idx}_{timestamp}.json"
            self.quality_eval_file = os.path.join(self.record_dir, eval_filename)
        
        # 确保记录目录存在
        os.makedirs(self.record_dir, exist_ok=True)
        
        # 初始化记录文件
        if not os.path.exists(self.record_file):
            with open(self.record_file, 'w', encoding='utf-8') as f:
                json.dump([], f, ensure_ascii=False, indent=2, default=str)
    
    def count_tokens(self, text, model="gpt-4o"):
        """计算文本的Token数量"""
        try:
            # 根据模型选择编码器
            if "gpt-4" in model.lower():
                encoding = tiktoken.encoding_for_model("gpt-4")
            elif "gpt-3.5" in model.lower():
                encoding = tiktoken.encoding_for_model("gpt-3.5-turbo")
            else:
                encoding = tiktoken.get_encoding("cl100k_base")
            
            return len(encoding.encode(str(text)))
        except Exception as e:
            print(f"❌ Token计算失败: {e}")
            # 粗略估算：中文约1个字符=1.5个token，英文约4个字符=1个token
            text_str = str(text)
            chinese_chars = len(re.findall(r'[\u4e00-\u9fff]', text_str))
            english_chars = len(text_str) - chinese_chars
            return int(chinese_chars * 1.5 + english_chars * 0.25)
    
    def record_conversion(self, data_index, original_table, tree_table, 
                         conversion_time, token_info, method, model_name, quality_metrics=None):
        """记录表格转换信息"""
        try:
            # 读取现有记录
            with open(self.record_file, 'r', encoding='utf-8') as f:
                records = json.load(f)
            
            # 创建新记录
            new_record = {
                "data_index": data_index,
                "timestamp": time.time(),
                "conversion_time_seconds": conversion_time,
                "method": method,
                "model_name": model_name,
                "token_info": token_info,
                "original_table_info": {
                    "rows": len(original_table),
                    "cols": len(original_table[0]) if original_table else 0,
                    "total_cells": len(original_table) * len(original_table[0]) if original_table else 0
                },
                "tree_table_info": {
                    "root_nodes_count": len(tree_table) if tree_table else 0,
                    "total_nodes": self._count_tree_nodes(tree_table) if tree_table else 0
                }
            }
            
            # 如果有质量评估指标，添加到记录中
            if quality_metrics:
                new_record["quality_metrics"] = {
                    "coverage_rate": quality_metrics["coverage"]["coverage_rate"],
                    "positioning_accuracy": quality_metrics["positioning"]["positioning_accuracy"],
                    "consistency_rate": quality_metrics["path_consistency"]["consistency_rate"],
                    "overall_score": quality_metrics["overall_score"]
                }
            
            # 添加新记录
            records.append(new_record)
            
            # 保存更新后的记录
            with open(self.record_file, 'w', encoding='utf-8') as f:
                json.dump(records, f, ensure_ascii=False, indent=2)
            
            print(f"✅ 转换记录已保存到: {self.record_file}")
            
        except Exception as e:
            print(f"❌ 保存转换记录失败: {e}")
    
    def _count_tree_nodes(self, tree_table):
        """递归计算树形表格的节点总数"""
        if not tree_table:
            return 0
        
        count = 0
        def count_nodes(node):
            nonlocal count
            count += 1
            if isinstance(node, dict) and 'children' in node:
                for child in node['children']:
                    count_nodes(child)
            elif isinstance(node, list):
                for item in node:
                    count_nodes(item)
        
        if isinstance(tree_table, list):
            for root in tree_table:
                count_nodes(root)
        else:
            count_nodes(tree_table)
        
        return count

from openpyxl import load_workbook
import os      
def load_xlsx_as_list(table_path):
    wb = load_workbook(table_path , data_only=True)
    ws = wb.active  # 默认读取第一个 sheet
    table_list = []
    for row in ws.iter_rows(values_only=True):
        table_list.append(list(row))
    return table_list

def convert_sheet_to_special_markdown(sheet: Worksheet) -> str:
    """
    将 openpyxl sheet 转换为特殊的 Markdown 表格。
    每个单元格的内容将是其地址（或合并范围）和其值。
    """
    
    # 一个集合，用于跟踪已被合并单元格覆盖的单元格坐标
    covered_cells = set()
    
    # 一个列表，用于存储 Markdown 表格每一行的数据（列表的列表）
    markdown_rows = []

    # 遍历工作表中的所有行和列
    for r_idx in range(1, sheet.max_row + 1):
        row_content = []
        for c_idx in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=r_idx, column=c_idx)
            coord = cell.coordinate

            # 1. 检查此单元格是否已被处理（作为合并单元格的一部分）
            if coord in covered_cells:
                row_content.append("")  # 在 Markdown 表格中留空
                continue

            # 2. 检查此单元格是否是新合并单元格的起始点
            is_merged = False
            for merged_range in sheet.merged_cells.ranges:
                if coord in merged_range:
                    is_merged = True
                    
                    # 获取合并范围的字符串表示，例如 "A2:B2"
                    range_str = str(merged_range)
                    
                    # 获取合并单元格的值（总是在左上角）
                    top_left_cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                    value = top_left_cell.value
                    
                    # 将此合并范围内的所有单元格添加到 'covered_cells' 集合
                    for mr in range(merged_range.min_row, merged_range.max_row + 1):
                        for mc in range(merged_range.min_col, merged_range.max_col + 1):
                            covered_cells.add(sheet.cell(row=mr, column=mc).coordinate)
                    
                    # 组合地址和值
                    value_str = "" if value is None else str(value)
                    cell_output = f"{range_str} {value_str}".strip()
                    row_content.append(cell_output)
                    
                    break # 找到了合并范围，跳出内部循环
            
            # 3. 如果它不是合并单元格（也不是被覆盖的单元格）
            if not is_merged:
                value = cell.value
                value_str = "" if value is None else str(value)
                cell_output = f"{coord} {value_str}".strip()
                row_content.append(cell_output)

        markdown_rows.append(row_content)

    # 4. 将 markdown_rows (列表的列表) 格式化为 Markdown 表格字符串
    if not markdown_rows:
        return ""

    # 创建 Markdown 表格
    md_output_lines = []
    
    # 添加表头
    header = markdown_rows[0]
    md_output_lines.append("| " + " | ".join(header) + " |")
    
    # 添加分隔符
    separator = ["---"] * len(header)
    md_output_lines.append("| " + " | ".join(separator) + " |")
    
    # 添加内容行
    for row in markdown_rows[1:]:
        md_output_lines.append("| " + " | ".join(row) + " |")

    return "\n".join(md_output_lines)
import pandas as pd
def sample_from_dataset(index, dataset_name="tatqa"):
    """
    从指定数据集加载样本数据
    
    Args:
        index: 数据索引
        dataset_name: 数据集名称 ("tatqa", "hitab", "aitqa", "sstqa", "tableeval")
    
    Returns:
        sample_data, whole_paragraph, table, questions, answers, table_uid
    """
    # 获取当前脚本所在目录，用于构建绝对路径
    dataset_dir = str(get_dataset_root())
    sheet = None
    if dataset_name.lower() == "tatqa":
        # TATQA
        tatqa_path = os.path.join(dataset_dir, "tatqa_dataset_train.json")
        tatqa_dataset = json.load(open(tatqa_path, "r"))
        sample_data = tatqa_dataset[index]
        table_uid = sample_data["table"]["uid"]
        paragraphs = sample_data["paragraphs"]
        whole_paragraph = [s["text"] for s in paragraphs]
        table = sample_data["table"]["table"]
        questions = [qa_pair["question"] for qa_pair in sample_data["questions"]]
        answers = [qa_pair["answer"] for qa_pair in sample_data["questions"]]
    
    elif dataset_name.lower() == "hitab":
        # HiTabQA
        hitab_dataset_path = os.path.join(dataset_dir, "hitab", "test_samples_clean.jsonl")
        hitab_dataset = [json.loads(line) for line in open(hitab_dataset_path, "r")]
        sample_data = hitab_dataset[index]
        
        # 适配HiTab数据格式（需要根据实际格式调整）
        table_uid = sample_data.get("table_id", f"hitab_{index}")
        whole_paragraph = []  # HiTab可能没有paragraph
        table_path = os.path.join(dataset_dir, "hitab", "tables", "raw", table_uid + ".json")
        table = json.load(open(table_path, "r"))
        table = table["texts"]
        questions = [sample_data.get("question", "")]
        answers = [str(sample_data.get("answer", ""))]
    
    elif dataset_name.lower() == "aitqa":
        # AIT-QA
        aitqa_questions_path = os.path.join(dataset_dir, "AIT-QA", "aitqa_clean_questions.json")
        aitqa_dataset = json.load(open(aitqa_questions_path, "r"))
        sample_data = aitqa_dataset[index]
        
        # 从AIT-QA数据格式提取信息
        table_uid = sample_data["table_id"]
        whole_paragraph = []  # AIT-QA没有paragraph字段
        table = sample_data["table"]
        questions = [sample_data["question"]]  # 单个问题转为列表
        answers = sample_data["answers"]  # answers已经是列表格式
    elif dataset_name.lower() == "tableeval":
        tableeval_path = os.path.join(dataset_dir, "TableEval", "TableEval-test.jsonl")
        tableeval_dataset = [json.loads(line) for line in open(tableeval_path, "r")]
        sample_data = tableeval_dataset[index]
        table_uid = sample_data["table_id"]
        whole_paragraph = []
        table = sample_data["context"]["context_markdown"]
        questions = [sample_data["golden_answer_list"][0]["问题列表"][0]["问题"]]
        answers = [sample_data["golden_answer_list"][0]["问题列表"][0]["最终答案"]]
    elif dataset_name.lower() == "sstqa":
        sstqa_path = os.path.join(dataset_dir, "SSTQA-zh", "test.jsonl")
        sstqa_dataset = [json.loads(line) for line in open(sstqa_path, "r")]
        sample_data = sstqa_dataset[index]
        table_uid = sample_data["table_id"]
        whole_paragraph = []
        questions = [sample_data["query"]]
        answers = [sample_data["label"]]
        table_path = os.path.join(dataset_dir, "SSTQA-zh", "table", str(table_uid) + ".xlsx")
        table = load_xlsx_as_list(table_path)
        sheet = load_workbook(table_path).active
    elif dataset_name.lower() == "realhitbench":
        realhitbench_path = os.path.join(dataset_dir, "RealHiTBench", "QA_final_filter.json")
        realhitbench_dataset = json.load(open(realhitbench_path, "r"))
        sample_data = realhitbench_dataset["queries"][index]
        table_uid = sample_data["FileName"]
        table_path = os.path.join(dataset_dir, "RealHiTBench", "csv", str(sample_data["FileName"]) + ".csv")
        table = pd.read_csv(table_path)
        whole_paragraph = []
        table = table.values.tolist()
        questions = [sample_data["Question"]]
        answers = [sample_data["ProcessedAnswer"]]
    elif dataset_name.lower() == "mmqa":
        mmqa_path = os.path.join(dataset_dir, "MMQA", "Synthesized_three_table.json")
        data = json.load(open(mmqa_path,"r"))
        sample_data = data[index]
        table_uid = sample_data["id_"]
        whole_paragraph = []
        questions = [sample_data["Question"]]
        answers = [sample_data["answer"]]
        tables_name = sample_data["table_names"]
        tables = sample_data["tables"]
        final_table = []
        for index in range(len(tables)):
            table = tables[index]
            table_name = tables_name[index]
            final_table.append("sub_table_name: " + table_name + "\n" + str(table))
        table = final_table
    else:
        raise ValueError(f"不支持的数据集: {dataset_name}，请使用 'tatqa', 'hitab' 或 'aitqa'")
    
    return sample_data, whole_paragraph, table, questions, answers, table_uid, sheet

def process_single_table(index, cache_manager, recorder, force_generate,
                        table2_tree_method, table2_tree_mode, model_name_table2tree, model_name_treeqa,
                        temperature_rule, model_type_treecons, model_type_qa, tree_qa,
                        force_generate_table_ids, dataset_name="tatqa", using_embedding=False):
    """处理单个表格的转换和问答
    
    Args:
        index: 数据索引
        cache_manager: 缓存管理器
        recorder: 记录器
        force_generate: 是否强制生成
        table2_tree_method: 表格转树的方法 ("rule" 或 "llm_based")
        model_name_table2tree: 表格转树阶段使用的模型名称
        model_name_treeqa: 问答阶段使用的模型名称
        temperature_rule: 温度参数
        model_type_treecons: 表格转树形结构阶段的模型类型 ("oai" 或 "opensource")
        model_type_qa: 问答阶段的模型类型 ("oai" 或 "opensource")
        dataset_name: 数据集名称
        using_embedding: 是否在TreeQA使用embedding mode 辅助检索
    """
    print(f"\n{'='*60}")
    print(f"🔢 处理数据索引: {index} (数据集: {dataset_name})")
    print(f"🔧 Tree构建模型: {model_type_treecons} ({model_name_table2tree}), QA模型: {model_type_qa} ({model_name_treeqa})")
    
    # 获取数据
    sample_data, whole_paragraph, table, questions, answers, table_uid, sheet = sample_from_dataset(index, dataset_name)
    
    # 记录开始时间
    start_time = time.time()
    
    # 第一步：检查缓存或生成树形表格
    print("🔍 检查是否存在缓存的树形表格...")
    Tree_table = None
    conversion_time = 0
    
    if cache_manager.has_cached_tree(table) and (not force_generate or table_uid in force_generate_table_ids):
        # 从缓存加载
        print("📦 发现缓存，正在加载...")
        Tree_table, row_h, col_h = cache_manager.load_tree_table(table)
        
        if Tree_table is not None:
            print(f"✅ 从缓存加载成功，根节点数量: {len(Tree_table)}")
        else:
            print("❌ 缓存加载失败，将重新生成...")
            Tree_table = None
    else:
        print("📝 未找到缓存或强制生成，开始生成树形表格...")
        Tree_table = None
    
    # 如果缓存加载失败或不存在，或强制生成，则重新生成
    if Tree_table is None :
        print("🔍 开始生成树形表格...")
        generation_start = time.time()
        force_generate_table_ids.append(table_uid)
        
        # 用于存储多次尝试的结果（树形表格、质量指标、覆盖率）
        attempts = []
        max_attempts = 3
        coverage_threshold = 0.9  # 信息覆盖率阈值 90%
        
        for attempt in range(1, max_attempts + 1):
            print(f"\n🔄 第 {attempt}/{max_attempts} 次生成尝试...")
            
            if table2_tree_method == "rule":
                current_tree = table2tree_rule(table, model_name_table2tree, temperature_rule, model_type_treecons)
            elif table2_tree_method == "llm_based":
                Tree_table_v1, Tree_table_v2 = table2tree_llm(sheet, table, model_name_table2tree, temperature_rule, model_type_treecons, table2_tree_mode)
                current_tree = Tree_table_v1
            
            print(f"✅ 生成完成，根节点数量: {len(current_tree)}")
            
            # 质量评估
            current_quality_metrics = None
            current_coverage_rate = 0.0
            if hasattr(recorder, 'enable_quality_eval') and recorder.enable_quality_eval:
                try:
                    print("\n🎯 开始树形结构质量评估...")
                    current_quality_metrics = evaluate_tree_quality(table, current_tree)
                    current_coverage_rate = current_quality_metrics.get("coverage", {}).get("coverage_rate", 0.0)
                    print(f"📊 信息覆盖率: {current_coverage_rate:.2%}")
                except Exception as e:
                    print(f"⚠️  质量评估失败: {e}")
                    current_quality_metrics = None
                    current_coverage_rate = 0.0
            
            # 保存当前尝试结果
            attempts.append({
                "tree": current_tree,
                "quality_metrics": current_quality_metrics,
                "coverage_rate": current_coverage_rate
            })
            
            # 如果覆盖率达到阈值，或者质量评估未启用，则停止尝试
            if current_coverage_rate >= coverage_threshold or current_coverage_rate <= 0.2:
                print(f"✨ 信息覆盖率 {current_coverage_rate:.2%} >= {coverage_threshold:.0%}，质量达标！")
                break
            elif not (hasattr(recorder, 'enable_quality_eval') and recorder.enable_quality_eval):
                print("ℹ️  质量评估未启用，使用当前生成结果")
                break
            elif attempt < max_attempts:
                print(f"⚠️  信息覆盖率 {current_coverage_rate:.2%} < {coverage_threshold:.0%}，将进行下一次尝试...")
            else:
                print(f"⚠️  已达到最大尝试次数 {max_attempts}，将选择覆盖率最高的版本")
        
        # 选择最佳版本（覆盖率最高的）
        if len(attempts) > 1:
            best_attempt = max(attempts, key=lambda x: x["coverage_rate"])
            best_idx = attempts.index(best_attempt) + 1
            if best_attempt["coverage_rate"] < coverage_threshold:
                print(f"\n🏆 选择第 {best_idx} 次尝试的结果（覆盖率最高: {best_attempt['coverage_rate']:.2%}）")
            Tree_table = best_attempt["tree"]
            quality_metrics = best_attempt["quality_metrics"]
        else:
            Tree_table = attempts[0]["tree"]
            quality_metrics = attempts[0]["quality_metrics"]
        
        conversion_time = time.time() - generation_start
        print(f"\n✅ 树形表格生成完成，根节点数量: {len(Tree_table)}, 总耗时: {conversion_time:.2f}秒")
        
        row_h = []
        col_h = []
        # 保存到缓存
        cache_manager.save_tree_table(table, Tree_table, row_h, col_h)
        
        # 计算Token信息
        original_table_text = json.dumps(table, ensure_ascii=False)
        tree_table_text = json.dumps(Tree_table, ensure_ascii=False)
        
        token_info = {
            "original_table_tokens": recorder.count_tokens(original_table_text, model_name_table2tree),
            "tree_table_tokens": recorder.count_tokens(tree_table_text, model_name_table2tree),
            "original_table_chars": len(original_table_text),
            "tree_table_chars": len(tree_table_text)
        }
        
        # 记录转换信息
        recorder.record_conversion(
            data_index=index,
            original_table=table,
            tree_table=Tree_table,
            conversion_time=conversion_time,
            token_info=token_info,
            method=table2_tree_method,
            model_name=model_name_table2tree,
            quality_metrics=quality_metrics
        )
    
    # 第二步：使用TreeQA处理所有问题
    all_results = []
    if questions and Tree_table:
        print(f"\n🤔 开始使用TreeQA处理 {len(questions)} 个问题...")
        
        
        for q_idx, (question, correct_answer) in enumerate(zip(questions, answers)):
            print(f"\n{'-'*40}")
            print(f"📋 问题 {q_idx + 1}/{len(questions)}: {question}")
            print(f"🎯 正确答案: {correct_answer}")
            
            try:
                # 处理问题并获取相关路径
                result, final_answer = tree_qa.process_question(Tree_table, table, question, whole_paragraph)
                
                symbolic_result = tree_qa.symbolic_tree_qa(Tree_table, table, question)
                if symbolic_result['error'] is not None:
                    print(f"❌ 符号化推理答案错误: {symbolic_result['error']}")
                    symbolic_answer = f"ERROR: {symbolic_result['error']}"
                else:
                    symbolic_answer = symbolic_result['answer']
                
                matches = re.findall(r'\[(.*?)\]', final_answer)
                if matches:
                    extract_answer = matches[-1]  # 取最后一个
                else:
                    extract_answer = ""
                    print("没有找到方括号内容")
                question_result = {
                    "question_index": q_idx,
                    "question": question,
                    "correct_answer": correct_answer,
                    "generated_answer": final_answer,
                    "generated_symbolic_code": symbolic_result['generated_code'],
                    "symbolic_answer": symbolic_answer,
                    "extra_answer": extract_answer,
                    "relevant_paths_count": len(result['relevant_paths']),
                    "path_details": result['path_details']
                }
                
                all_results.append(question_result)
                
                print(f"🎯 正确答案: {correct_answer}\n✅ 生成答案: {final_answer}\n✅符号化推理答案: {symbolic_answer}")
                
            except Exception as e:
                print(f"❌ 处理问题时出错: {e}")
                error_result = {
                    "question_index": q_idx,
                    "question": question,
                    "correct_answer": correct_answer,
                    "generated_answer": f"ERROR: {str(e)}",
                    "relevant_paths_count": 0,
                    "path_details": []
                }
                all_results.append(error_result)
    
    total_time = time.time() - start_time
    print(f"\n✅ 表格 {index} 处理完成，总耗时: {total_time:.2f}秒")
    
    return {
        "data_index": index,
        "table_uid": table_uid,
        "total_time": total_time,
        "conversion_time": conversion_time,
        "questions_count": len(questions),
        "results": all_results
    }


def build_arg_parser():
    parser = argparse.ArgumentParser(
        description="Run the ASTRA table QA pipeline over a dataset split."
    )
    parser.add_argument(
        "--dataset",
        default="hitab",
        help="Dataset name, e.g. hitab, tatqa, aitqa, sstqa, realhitbench, mmqa.",
    )
    parser.add_argument(
        "--table2-tree-method",
        default="llm_based",
        choices=["rule", "llm_based"],
        help="Tree construction method.",
    )
    parser.add_argument(
        "--table2-tree-mode",
        default="normal",
        choices=["normal", "enhanced"],
        help="Prompting mode for LLM-based tree construction.",
    )
    parser.add_argument(
        "--model-name-table2tree",
        default="gpt-4o",
        help="Model name for the table-to-tree stage.",
    )
    parser.add_argument(
        "--model-name-treeqa",
        default="gpt-4o",
        help="Model name for the QA stage.",
    )
    parser.add_argument(
        "--model-type-treecons",
        default="oai",
        choices=["oai", "opensource"],
        help="Backend type for the table-to-tree stage.",
    )
    parser.add_argument(
        "--model-type-qa",
        default="oai",
        choices=["oai", "opensource"],
        help="Backend type for the QA stage.",
    )
    parser.add_argument(
        "--start-index",
        type=int,
        default=0,
        help="Inclusive start index.",
    )
    parser.add_argument(
        "--end-index",
        type=int,
        default=100,
        help="Inclusive end index.",
    )
    parser.add_argument(
        "--temperature-rule",
        type=float,
        default=0.3,
        help="Temperature used during tree construction.",
    )
    parser.add_argument(
        "--force-generate",
        action="store_true",
        help="Ignore cached trees and regenerate them.",
    )
    parser.add_argument(
        "--using-embedding",
        dest="using_embedding",
        action="store_true",
        help="Enable embedding-assisted retrieval in TreeQA.",
    )
    parser.add_argument(
        "--no-embedding",
        dest="using_embedding",
        action="store_false",
        help="Disable embedding-assisted retrieval in TreeQA.",
    )
    parser.set_defaults(using_embedding=True)
    parser.add_argument(
        "--enable-quality-eval",
        dest="enable_quality_eval",
        action="store_true",
        help="Run tree quality evaluation during tree construction.",
    )
    parser.add_argument(
        "--disable-quality-eval",
        dest="enable_quality_eval",
        action="store_false",
        help="Skip tree quality evaluation for faster runs.",
    )
    parser.set_defaults(enable_quality_eval=True)
    return parser


def main():
    args = build_arg_parser().parse_args()

    if args.end_index < args.start_index:
        raise ValueError("--end-index must be greater than or equal to --start-index")

    cache_manager = TreeTableCache()
    recorder = Table2TreeRecorder(
        dataset_name=args.dataset,
        start_idx=args.start_index,
        end_idx=args.end_index,
        enable_quality_eval=args.enable_quality_eval
    )

    print(f"🚌 开始批量处理数据，范围: {args.start_index} - {args.end_index}")

    batch_start_time = time.time()
    all_table_results = []
    force_generate_table_ids = []
    tree_qa = TreeQA(
        model_type=args.model_type_qa,
        model_name=args.model_name_treeqa,
        using_embedding=args.using_embedding
    )

    for index in range(args.start_index, args.end_index + 1):
        try:
            table_result = process_single_table(
                index=index,
                cache_manager=cache_manager,
                recorder=recorder,
                force_generate=args.force_generate,
                table2_tree_method=args.table2_tree_method,
                table2_tree_mode=args.table2_tree_mode,
                model_name_table2tree=args.model_name_table2tree,
                model_name_treeqa=args.model_name_treeqa,
                temperature_rule=args.temperature_rule,
                model_type_treecons=args.model_type_treecons,
                model_type_qa=args.model_type_qa,
                tree_qa=tree_qa,
                force_generate_table_ids=force_generate_table_ids,
                dataset_name=args.dataset,
                using_embedding=args.using_embedding
            )
            all_table_results.append(table_result)
        except Exception as e:
            print(f"❓ 处理表格 {index} 时报错: {e}")
            error_result = {
                "data_index": index,
                "total_time": 0,
                "conversion_time": 0,
                "questions_count": 0,
                "results": [],
                "error": str(e)
            }
            all_table_results.append(error_result)

    batch_total_time = time.time() - batch_start_time
    print(f"\n{'='*60}")
    print("🎀 批量处理完成！")
    print("📳 总结:")
    print(f"   - 处理表格数量: {len(all_table_results)}")
    print(f"   - 总耗时: {batch_total_time:.2f}秒")

    if all_table_results:
        print(f"   - 平均每表格耗时: {batch_total_time / len(all_table_results):.2f}秒")

    total_questions = sum(result["questions_count"] for result in all_table_results)
    successful_tables = len([r for r in all_table_results if "error" not in r])

    print(f"   - 成功处理表格: {successful_tables}/{len(all_table_results)}")
    print(f"   - 总问题数量: {total_questions}")
    print(f"   - 记录文件: {recorder.record_file}")

    batch_result_file = recorder.batch_result_file
    try:
        with open(batch_result_file, "w", encoding="utf-8") as f:
            json.dump({
                "batch_info": {
                    "start_index": args.start_index,
                    "end_index": args.end_index,
                    "total_time": batch_total_time,
                    "processed_tables": len(all_table_results),
                    "successful_tables": successful_tables,
                    "total_questions": total_questions
                },
                "table_results": all_table_results
            }, f, ensure_ascii=False, indent=2)
        print(f"   - 批量结果已保存: {batch_result_file}")
    except Exception as e:
        print(f"❓ 保存批量结果失败: {e}")

if __name__ == "__main__":
    main()
