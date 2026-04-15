"""
Pipeline wrapper for ASTRA demo - exposes stepwise interfaces for frontend visualization.
"""
import sys
import os
import json
import io
import re
import tempfile

# Add parent directory to path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from astra_config import get_default_embedding_path
from table2tree import llm_based_converter, table2tree_rule, normalize_tree_none_values, excel_to_markdown_with_cell_ref, convert_cell_positions_to_values
from treeqa import TreeQA, extract_all_nodes
from model_clients import ModelClient, OpenaimodelClient


def preprocess_upload(file_bytes: bytes, filename: str):
    """Convert uploaded file to unified list-of-lists format.
    
    Returns: (table, sheet) where sheet is openpyxl worksheet or None
    """
    ext = os.path.splitext(filename)[1].lower()
    sheet = None
    
    if ext == '.json':
        content = file_bytes.decode('utf-8')
        data = json.loads(content)
        if isinstance(data, list) and len(data) > 0:
            if isinstance(data[0], list):
                table = data  # already list of lists
            elif isinstance(data[0], dict):
                # list of dicts -> list of lists
                headers = list(data[0].keys())
                table = [headers] + [[row.get(h, '') for h in headers] for row in data]
            else:
                table = data
        else:
            table = data
            
    elif ext == '.xlsx':
        from openpyxl import load_workbook
        # Save to temp file for openpyxl
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp.write(file_bytes)
            tmp_path = tmp.name
        wb = load_workbook(tmp_path, data_only=True)
        sheet = wb.active
        table = []
        for row in sheet.iter_rows(values_only=True):
            table.append([v if v is not None else '' for v in list(row)])
        os.unlink(tmp_path)
        
    elif ext == '.csv':
        import pandas as pd
        content = file_bytes.decode('utf-8')
        df = pd.read_csv(io.StringIO(content))
        table = [df.columns.tolist()] + df.values.tolist()
        
    elif ext == '.md':
        content = file_bytes.decode('utf-8')
        table = parse_markdown_table(content)
        
    else:
        raise ValueError(f"Unsupported file format: {ext}")
    
    # Clean None values
    cleaned_table = []
    for row in table:
        cleaned_row = []
        for cell in row:
            if cell is None:
                cleaned_row.append('')
            else:
                cleaned_row.append(cell)
        cleaned_table.append(cleaned_row)
    
    return cleaned_table, sheet


def parse_markdown_table(md_text: str):
    """Parse a markdown table into list-of-lists."""
    lines = [l.strip() for l in md_text.strip().split('\n') if l.strip()]
    table = []
    for line in lines:
        if line.startswith('|') and not all(c in '|-: ' for c in line):
            cells = [c.strip() for c in line.split('|')[1:-1]]
            table.append(cells)
    return table if table else [[md_text]]


def build_tree_stepwise(table, sheet, model_name, temperature, model_type, mode="normal",
                        api_key="", base_url=""):
    """Generator that yields intermediate results at each tree construction stage.
    
    Yields dicts with keys: stage, status, data
    """
    converter = llm_based_converter(model_type, model_name, api_key=api_key, base_url=base_url)
    
    try:
        if mode == "normal":
            # Stage 1: Header Normalization
            yield {"stage": "header_normalization", "status": "running", "data": None}
            normalized_headers = converter.header_analysis_norm(table, temperature)
            yield {
                "stage": "header_normalization", 
                "status": "completed",
                "data": {"raw_output": normalized_headers}
            }
            
            # Stage 2: Hierarchy Identification
            yield {"stage": "hierarchy_identification", "status": "running", "data": None}
            hierarchy_definition = converter.hierarchy_definition(table, normalized_headers, temperature)
            yield {
                "stage": "hierarchy_identification",
                "status": "completed", 
                "data": {"raw_output": hierarchy_definition}
            }
            
            # Stage 3: Tree Construction
            yield {"stage": "tree_construction", "status": "running", "data": None}
            final_json_tree = converter.final_json_tree_construction(table, normalized_headers, hierarchy_definition, temperature)
            
            if model_type == "oai":
                tree_table = converter.parse_json_with_merge_oai(final_json_tree)
            else:
                tree_table = converter.parse_json_with_merge(final_json_tree)
            
            yield {
                "stage": "tree_construction",
                "status": "completed",
                "data": {
                    "raw_output": final_json_tree,
                    "tree_table": tree_table
                }
            }
            
        elif mode == "enhanced":
            if sheet is None:
                raise ValueError("Enhanced mode requires an xlsx sheet")
            markdown_table = excel_to_markdown_with_cell_ref(sheet)
            
            yield {"stage": "header_normalization", "status": "running", "data": {"markdown_table": markdown_table}}
            normalized_headers = converter.header_analysis_norm_enhance(markdown_table, temperature)
            yield {
                "stage": "header_normalization",
                "status": "completed",
                "data": {"raw_output": normalized_headers, "markdown_table": markdown_table}
            }
            
            yield {"stage": "hierarchy_identification", "status": "running", "data": None}
            hierarchy_definition = converter.hierarchy_definition_enhance(markdown_table, normalized_headers, temperature)
            yield {
                "stage": "hierarchy_identification",
                "status": "completed",
                "data": {"raw_output": hierarchy_definition}
            }
            
            yield {"stage": "tree_construction", "status": "running", "data": None}
            final_json_tree = converter.final_json_tree_construction_enhance(markdown_table, normalized_headers, hierarchy_definition, temperature)
            
            if model_type == "oai":
                tree_table = converter.parse_json_with_merge_oai(final_json_tree)
            else:
                tree_table = converter.parse_json_with_merge(final_json_tree)
            
            tree_table = convert_cell_positions_to_values(tree_table, sheet)
            
            yield {
                "stage": "tree_construction",
                "status": "completed",
                "data": {
                    "raw_output": final_json_tree,
                    "tree_table": tree_table
                }
            }
            
    except Exception as e:
        yield {"stage": "error", "status": "failed", "data": {"error": str(e)}}


class DemoTreeQA:
    """Wrapper around TreeQA that exposes stepwise reasoning for the demo."""
    
    def __init__(self, model_type="oai", model_name="deepseek-v3-250324",
                 api_key="", base_url="",
                 using_embedding=False, embedding_path=None,
                 embedding_model_name="", embedding_api_key="", embedding_base_url=""):
        self.tree_qa = TreeQA(
            model_type=model_type,
            model_name=model_name,
            api_key=api_key,
            base_url=base_url,
            using_embedding=using_embedding,
            embedding_path=embedding_path or get_default_embedding_path(),
            embedding_model_name=embedding_model_name,
            embedding_api_key=embedding_api_key,
            embedding_base_url=embedding_base_url,
        )
    
    def qa_stepwise(self, tree_table, original_table, question, whole_paragraph=None):
        """Generator that yields intermediate QA reasoning steps.
        
        Yields dicts with: stage, status, data
        """
        if whole_paragraph is None:
            whole_paragraph = []
        
        tree_table = normalize_tree_none_values(tree_table)
        
        # Stage 1: Embedding retrieval (optional)
        embedding_path_list = []
        if self.tree_qa.using_embedding:
            yield {"stage": "embedding_retrieval", "status": "running", "data": None}
            try:
                top_k_results = self.tree_qa.embedding_revelant_paths(tree_table, original_table, question)
                embedding_path_list = [item[0][0] for item in top_k_results]
                self.tree_qa.embedding_re_path = embedding_path_list
                yield {
                    "stage": "embedding_retrieval",
                    "status": "completed",
                    "data": {
                        "top_k_nodes": [
                            {"path": item[0][0], "key": str(item[0][1]), "score": round(item[1], 2)}
                            for item in top_k_results
                        ]
                    }
                }
            except Exception as e:
                yield {"stage": "embedding_retrieval", "status": "failed", "data": {"error": str(e)}}
        
        # Stage 2: Path planning guide
        yield {"stage": "plan_path_guide", "status": "running", "data": None}
        try:
            llm_guide_path = self.tree_qa.plan_path_guide(tree_table, question)
            yield {
                "stage": "plan_path_guide",
                "status": "completed",
                "data": {"guide_text": llm_guide_path}
            }
        except Exception as e:
            llm_guide_path = ""
            yield {"stage": "plan_path_guide", "status": "failed", "data": {"error": str(e)}}
        
        # Stage 3: Adaptive Tree Navigation (stepwise DFS)
        yield {"stage": "tree_navigation", "status": "running", "data": None}
        navigation_steps = []
        relevant_paths = []
        
        stack = [("ROOT", tree_table, ["ROOT"])]
        step_count = 0
        
        while stack:
            current_node_name, current_data, current_path = stack.pop()
            step_count += 1
            
            if isinstance(current_data, dict) and any(isinstance(v, dict) for v in current_data.values()):
                child_nodes = {k: v for k, v in current_data.items()}
                if child_nodes:
                    child_names = list(child_nodes.keys())
                    selected_children = self.tree_qa._select_relevant_children(
                        current_node_name, child_nodes, question, current_path, 
                        embedding_path_list, llm_guide_path
                    )
                    
                    nav_step = {
                        "step": step_count,
                        "current_node": current_node_name,
                        "current_path": current_path,
                        "available_children": child_names,
                        "selected_children": selected_children,
                        "rejected_children": [c for c in child_names if c not in selected_children]
                    }
                    navigation_steps.append(nav_step)
                    
                    yield {
                        "stage": "tree_navigation_step",
                        "status": "in_progress",
                        "data": nav_step
                    }
                    
                    for child_name in selected_children:
                        if child_name in child_nodes:
                            new_path = current_path + [child_name]
                            stack.append((child_name, child_nodes[child_name], new_path))
            
            elif isinstance(current_data, dict) and all(not isinstance(v, dict) for v in current_data.values()):
                relevant_paths.append(current_path)
            elif isinstance(current_data, (str, int, float)) or current_data is None:
                relevant_paths.append(current_path)
        
        yield {
            "stage": "tree_navigation",
            "status": "completed",
            "data": {
                "total_steps": step_count,
                "navigation_steps": navigation_steps,
                "relevant_paths": relevant_paths
            }
        }
        
        # Stage 4: Collect evidence paths
        path_details = []
        for path in relevant_paths:
            data = self.tree_qa._get_path_data(tree_table, path)
            path_details.append({"path": path, "data": data})
        
        yield {
            "stage": "evidence_paths",
            "status": "completed",
            "data": {
                "paths": relevant_paths,
                "path_details": path_details
            }
        }
        
        # Stage 5: Generate final answer
        yield {"stage": "final_answer", "status": "running", "data": None}
        final_paths = []
        for pd_item in path_details:
            path_str = ' -> '.join(pd_item['path'])
            data_str = json.dumps(pd_item['data'], ensure_ascii=False, indent=2, default=str)
            final_paths.append(f"{path_str} : {data_str}")
        
        try:
            final_answer = self.tree_qa.get_final_answer(whole_paragraph, question, final_paths)
            yield {
                "stage": "final_answer",
                "status": "completed",
                "data": {"answer": final_answer, "evidence_paths": final_paths}
            }
        except Exception as e:
            yield {"stage": "final_answer", "status": "failed", "data": {"error": str(e)}}
    
    def symbolic_qa_stepwise(self, tree_table, question, max_iter_num=3):
        """Generator that yields intermediate symbolic reasoning steps."""
        tree_str = json.dumps(tree_table, indent=2, ensure_ascii=False)
        
        yield {"stage": "symbolic_reasoning", "status": "running", "data": None}
        
        last_error = None
        last_code = None
        
        for attempt in range(max_iter_num):
            yield {
                "stage": "symbolic_attempt",
                "status": "running",
                "data": {"attempt": attempt + 1, "max_attempts": max_iter_num}
            }
            
            try:
                if attempt == 0:
                    prompt = self.tree_qa._create_symbolic_tree_prompt(tree_str, question)
                else:
                    prompt = self.tree_qa._create_retry_prompt(tree_str, question, last_code, last_error, attempt)
                
                if self.tree_qa.model_type == "oai":
                    response = self.tree_qa.oaimodel_client.generate(prompt, temperature=0.7)
                else:
                    response = self.tree_qa.model_client.test_generate_stream(prompt, temperature=0.7)
                
                match = re.search(r"<python_code>(.*?)</python_code>", response, re.DOTALL | re.IGNORECASE)
                if not match:
                    last_error = "LLM did not return code in <python_code> tags."
                    last_code = response
                    yield {
                        "stage": "symbolic_attempt",
                        "status": "failed",
                        "data": {"attempt": attempt + 1, "error": last_error, "raw_response": response[:500]}
                    }
                    continue
                
                extracted_code = match.group(1).strip()
                extracted_code = "\n".join([line for line in extracted_code.splitlines() if not line.strip().startswith('#')])
                
                if not extracted_code:
                    last_error = "Empty code block."
                    last_code = ""
                    continue
                
                answer = self.tree_qa._safe_eval(extracted_code, tree_table)
                
                yield {
                    "stage": "symbolic_reasoning",
                    "status": "completed",
                    "data": {
                        "answer": answer,
                        "generated_code": extracted_code,
                        "attempts": attempt + 1
                    }
                }
                return
                
            except Exception as e:
                last_error = str(e)
                last_code = locals().get('extracted_code', None)
                yield {
                    "stage": "symbolic_attempt",
                    "status": "failed",
                    "data": {"attempt": attempt + 1, "error": last_error, "code": last_code}
                }
        
        yield {
            "stage": "symbolic_reasoning",
            "status": "failed",
            "data": {"error": f"Failed after {max_iter_num} attempts: {last_error}", "last_code": last_code}
        }
