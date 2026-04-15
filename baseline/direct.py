"""
Direct Baseline for Table QA
直接将问题和表格一起发送给模型进行回答，作为基线方法
"""

import json
import os
import sys
import time
import re
from datetime import datetime

# 添加父目录到路径以导入相关模块
current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.dirname(current_dir)
sys.path.append(parent_dir)

from astra_config import get_dataset_root, get_local_model_base_url
from model_clients import ModelClient, OpenaimodelClient
from tableqa import sample_from_dataset as shared_sample_from_dataset

def direct_answer(question: str, table: str, client, model_type="opensource"):
    """
    直接基于表格回答问题
    
    Args:
        question: 问题
        table: 表格数据（JSON字符串或列表）
        client: 模型客户端
        model_type: 模型类型 ("opensource" 或 "oai")
    
    Returns:
        生成的答案文本
    """
    # 将表格转为字符串格式
    if isinstance(table, str):
        table_str = table
    else:
        table_str = json.dumps(table, ensure_ascii=False, indent=2, default=str)
    
    prompt = f"""
你是一名高水平的表格分析专家。你的任务是基于以下表格回答问题。

# 注意事项：

1. 请将最终答案用方括号括起来，例如：[答案]。
2. **务必注意问题所要求的答案类型。**
    - 如果问题以 **"Which"** 或 **"What"** 开头（例如："Which item?"、"What category?"），你的答案必须是*名称或描述*（例如："Sell Product A"），**而不是**与其关联的数值。
    - 如果问题以 **"How much"**、**"What is the value"** 或 **"How many"** 开头，则答案必须是*数值*（例如："120000"）。
3. **仅当**答案是数值类型（根据规则2判断）时，你必须检查是否存在相关的"单位"（例如："万"、"百万"），并提供最终经过换算的数值结果。

## 表格:
{table_str}

## 问题:
{question}

## 最终答案:
"""
    
    try:
        if model_type == "opensource":
            response = client.test_generate_stream(prompt, max_length=16384, temperature=0.3)
            return response

        else:  # oai
            response = client.test_generate_stream(prompt, max_length=16384, temperature=0.3)
            return response
    except Exception as e:
        print(f"❌ 生成答案时出错: {e}")
        return f"ERROR: {str(e)}"


def sample_from_dataset(index, dataset_name="tatqa"):
    """
    从指定数据集加载样本数据
    
    Args:
        index: 数据索引
        dataset_name: 数据集名称
    
    Returns:
        sample_data, whole_paragraph, table, questions, answers, table_uid
    """
    # 获取dataset目录的绝对路径
    project_root = os.path.dirname(os.path.dirname(parent_dir))
    dataset_dir = str(get_dataset_root())
    
    if dataset_name.lower() == "tatqa":
        tatqa_path = os.path.join(dataset_dir, "tatqa_dataset_train.json")
        tatqa_dataset = json.load(open(tatqa_path, "r"))
        sample_data = tatqa_dataset[index]
        table_uid = sample_data["table"]["uid"]
        paragraphs = sample_data["paragraphs"]
        whole_paragraph = [s["text"] for s in paragraphs]
        table = sample_data["table"]["table"]
        questions = [qa_pair["question"] for qa_pair in sample_data["questions"]]
        answers = [qa_pair["answer"] for qa_pair in sample_data["questions"]]
    
    elif dataset_name.lower() == "hitab":
        hitab_dataset_path = os.path.join(dataset_dir, "hitab", "train_samples.jsonl")
        hitab_dataset = [json.loads(line) for line in open(hitab_dataset_path, "r")]
        sample_data = hitab_dataset[index]
        
        table_uid = sample_data.get("table_id", f"hitab_{index}")
        whole_paragraph = []
        table_path = os.path.join(dataset_dir, "hitab", "tables", "raw", table_uid + ".json")
        table = json.load(open(table_path, "r"))
        table = table["texts"]
        questions = [sample_data.get("question", "")]
        answers = [str(sample_data.get("answer", ""))]
    
    elif dataset_name.lower() == "aitqa":
        aitqa_questions_path = os.path.join(dataset_dir, "AIT-QA", "aitqa_clean_questions.json")
        aitqa_dataset = json.load(open(aitqa_questions_path, "r"))
        sample_data = aitqa_dataset[index]
        
        table_uid = sample_data["table_id"]
        whole_paragraph = []
        table = sample_data["table"]
        questions = [sample_data["question"]]
        answers = sample_data["answers"]
    
    elif dataset_name.lower() == "sstqa":
        from openpyxl import load_workbook
        sstqa_path = os.path.join(dataset_dir, "SSTQA-zh", "test.jsonl")
        sstqa_dataset = [json.loads(line) for line in open(sstqa_path, "r")]
        sample_data = sstqa_dataset[index]
        table_uid = sample_data["table_id"]
        whole_paragraph = []
        questions = [sample_data["query"]]
        answers = [sample_data["label"]]
        table_path = os.path.join(dataset_dir, "SSTQA-zh", "table", str(table_uid) + ".xlsx")
        
        # 加载xlsx表格
        wb = load_workbook(table_path, data_only=True)
        ws = wb.active
        table = []
        for row in ws.iter_rows(values_only=True):
            table.append(list(row))
    
    elif dataset_name.lower() == "tableeval":
        tableeval_path = os.path.join(dataset_dir, "TableEval", "TableEval-test.jsonl")
        tableeval_dataset = [json.loads(line) for line in open(tableeval_path, "r")]
        sample_data = tableeval_dataset[index]
        table_uid = sample_data["table_id"]
        whole_paragraph = []
        table = sample_data["context"]["context_markdown"]
        questions = [sample_data["golden_answer_list"][0]["问题列表"][0]["问题"]]
        answers = [sample_data["golden_answer_list"][0]["问题列表"][0]["最终答案"]]
    
    else:
        raise ValueError(f"不支持的数据集: {dataset_name}")
    
    return sample_data, whole_paragraph, table, questions, answers, table_uid


def process_single_sample(index, dataset_name, client, model_type):
    """处理单个样本"""
    print(f"\n{'='*60}")
    print(f"🔢 处理数据索引: {index} (数据集: {dataset_name})")
    
    try:
        # 加载数据
        sample_data, whole_paragraph, table, questions, answers, table_uid, _ = shared_sample_from_dataset(
            index, dataset_name
        )
        
        results = []
        start_time = time.time()
        
        # 处理每个问题
        for q_idx, (question, correct_answer) in enumerate(zip(questions, answers)):
            print(f"\n{'-'*40}")
            print(f"📋 问题 {q_idx + 1}/{len(questions)}: {question}")
            print(f"🎯 正确答案: {correct_answer}")
            
            # 生成答案
            question_start = time.time()
            generated_answer = direct_answer(question, table, client, model_type)
            question_time = time.time() - question_start
            
            # 提取方括号内的答案
            matches = re.findall(r'\[(.*?)\]', generated_answer)
            extracted_answer = matches[-1] if matches else ""
            
            print(f"✅ 生成答案: {generated_answer}")
            print(f"📦 提取答案: {extracted_answer}")
            print(f"⏱️  耗时: {question_time:.2f}秒")
            
            results.append({
                "question_index": q_idx,
                "question": question,
                "correct_answer": correct_answer,
                "generated_answer": generated_answer,
                "extracted_answer": extracted_answer,
                "generation_time": question_time
            })
        
        total_time = time.time() - start_time
        print(f"\n✅ 样本 {index} 处理完成，总耗时: {total_time:.2f}秒")
        
        return {
            "data_index": index,
            "table_uid": table_uid,
            "total_time": total_time,
            "questions_count": len(questions),
            "results": results
        }
        
    except Exception as e:
        print(f"❌ 处理样本 {index} 时出错: {e}")
        return {
            "data_index": index,
            "error": str(e),
            "table_results": []
        }


def main():
    """主函数"""
    # ==================== 配置参数 ====================
    dataset_name = "aitqa"  # 数据集: tatqa, hitab, aitqa, sstqa, tableeval
    model_type = "opensource"  # "opensource" 或 "oai"
    model_name = "qwen3-8b-instruct"  # 模型名称
    
    # 处理范围
    start_index = 0
    end_index = 386
    
    # ==================== 初始化 ====================
    print(f"🚀 Direct Baseline 开始运行")
    print(f"📊 数据集: {dataset_name}")
    print(f"🤖 模型类型: {model_type} ({model_name})")
    print(f"📈 处理范围: {start_index} - {end_index}")
    
    # 初始化模型客户端
    if model_type == "opensource":
        client = ModelClient(base_url=get_local_model_base_url("http://localhost:8002"))
    else:
        client = OpenaimodelClient(model=model_name)
    
    # 创建结果保存目录
    record_dir = os.path.join(parent_dir, "record")
    os.makedirs(record_dir, exist_ok=True)
    
    # 生成结果文件名
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    result_file = os.path.join(
        record_dir, 
        f"direct_baseline_{dataset_name}_{start_index}-{end_index}_{timestamp}.json"
    )
    
    # ==================== 批量处理 ====================
    all_results = []
    batch_start_time = time.time()
    
    for index in range(start_index, end_index + 1):
        result = process_single_sample(index, dataset_name, client, model_type)
        all_results.append(result)
    
    batch_total_time = time.time() - batch_start_time
    
    # ==================== 保存结果 ====================
    output_data = {
        "method": "direct_baseline",
        "dataset": dataset_name,
        "model_type": model_type,
        "model_name": model_name,
        "batch_info": {
            "start_index": start_index,
            "end_index": end_index,
            "total_time": batch_total_time,
            "processed_samples": len(all_results),
            "total_questions": sum(r['questions_count'] for r in all_results if 'questions_count' in r)
        },
        "table_results": all_results
    }
    
    with open(result_file, 'w', encoding='utf-8') as f:
        json.dump(output_data, f, ensure_ascii=False, indent=2)
    
    # ==================== 输出总结 ====================
    print(f"\n{'='*60}")
    print("🎉 批量处理完成！")
    print(f"📊 总结:")
    print(f"   - 处理样本数量: {len(all_results)}")
    print(f"   - 总耗时: {batch_total_time:.2f}秒")
    print(f"   - 平均每样本耗时: {batch_total_time/len(all_results):.2f}秒")
    print(f"   - 结果已保存: {result_file}")


if __name__ == "__main__":
    main()
